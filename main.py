import argparse
import os
import re
from fuzzysearch import find_near_matches
from openpyxl import load_workbook


def search_files(search_term, path):
    results = []

    for root, dirs, files in os.walk(path):
        for file in files:
            if file.lower().endswith(('.txt', '.xlsx', '.pdf')):
                file_path = os.path.join(root, file)

                if file.lower().endswith('.pdf'):
                    results.extend(search_pdf(file_path, search_term))
                elif file.lower().endswith('.xlsx'):
                    results.extend(search_excel(file_path, search_term))
                else:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                        matches = find_near_matches(search_term, content, max_l_dist=2)
                        if matches:
                            results.append((file_path, matches))
    return results


def search_pdf(file_path, search_term):
    # Logik für die Suche in PDF-Dateien (hier wird pypdf verwendet)
    import pypdf

    results = []
    with open(file_path, 'rb') as pdf_file:
        pdf_reader = pypdf.PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            matches = find_near_matches(search_term, text, max_l_dist=2)
            if matches:
                results.append((file_path, page_num + 1, matches))

    return results


def search_excel(file_path, search_term):
    results = []
    wb = load_workbook(file_path, read_only=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    matches = find_near_matches(search_term, cell.value, max_l_dist=2)
                    if matches:
                        results.append((file_path, sheet_name, cell.coordinate, matches))

    return results


def format_results(results):
    if not results:
        print("Keine Übereinstimmungen gefunden.")
        return

    print("Ergebnisse:")
    for file_path, *matches in results:
        print(f"Datei: {file_path}")
        for match in matches:
            print(f"  Seite: {match[1] if len(match) == 3 else 'N/A'} - Übereinstimmung: '{match[0]}'")


def main():
    parser = argparse.ArgumentParser(description="Sucht nach einem Suchbegriff in txt, xlsx, und pdf Dateien.")
    parser.add_argument("-s", "--search", required=True, help="Suchbegriff")
    parser.add_argument("-p", "--path", required=True, help="Dateipfad")

    args = parser.parse_args()

    results = search_files(args.search, args.path)
    format_results(results)


if __name__ == "__main__":
    main()
